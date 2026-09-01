"""
Excel Exporter for API 5L PSL2 & BOTAŞ Pipe QA/QC Suite.
Generates fully formatted, colored, print-ready .xlsx workbooks matching exact engineering layout.
"""

import io
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    @staticmethod
    def export_matrix_to_excel(
        project_info: Dict[str, Any],
        pipes_data: List[Dict[str, Any]],
        lang: str = "tr"
    ) -> io.BytesIO:
        """
        Exports the entire multi-pipe QA/QC inspection matrix to an Excel workbook.
        Populates all 40+ row standard remarks in Turkish or English.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Boru Kontrol ve Kabul Matrisi"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        font_title = Font(name="Arial", size=14, bold=True, color="1E3A8A")
        font_subtitle = Font(name="Arial", size=9, italic=True, color="475569")
        font_header_dark = Font(name="Arial", size=9, bold=True, color="0F172A")
        font_header_bold = Font(name="Arial", size=9, bold=True, color="000000")
        font_regular = Font(name="Arial", size=9, bold=False, color="000000")
        font_sub = Font(name="Arial", size=8, italic=True, color="475569")

        # Fills
        fill_header_main = PatternFill(start_color="9CA3AF", end_color="9CA3AF", fill_type="solid")
        fill_header_sub = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")
        fill_section = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_disclaimer = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        # Borders
        border_thin = Side(border_style="thin", color="94A3B8")
        border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

        # Alignments
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        num_pipes = len(pipes_data)
        max_col = max(7, 2 + num_pipes)
        last_col_letter = get_column_letter(max_col)

        # 1. Project Title & Metadata Header
        ws.merge_cells(f"A1:{last_col_letter}1")
        title_text = "API 5L PSL2 & BOTAŞ BORU KALİTE GÜVENCE VE KABUL MATRİSİ" if lang == "tr" else "API 5L PSL2 & BOTAŞ PIPE QA/QC ACCEPTANCE MATRIX"
        cell_t = ws.cell(1, 1, title_text)
        cell_t.font = font_title
        cell_t.alignment = align_left

        ws.merge_cells(f"A2:{last_col_letter}2")
        p_name = project_info.get('project_name', 'Boru Hattı Projesi')
        p_no = project_info.get('project_no', 'PRJ-2026')
        p_rev = project_info.get('revision', 'Rev. 0')
        sub_text = f"Proje: {p_name} | No: {p_no} | Revizyon: {p_rev} | Standart: {project_info.get('standard', 'BOTAŞ Şartnamesi')}"
        cell_s = ws.cell(2, 1, sub_text)
        cell_s.font = font_subtitle
        cell_s.alignment = align_left

        start_row = 5
        num_pipes = len(pipes_data)
        exp_dict = pipes_data[0].get('explanations', {}) if pipes_data else {}

        def get_exp(key: str, fallback: str = "") -> str:
            val = exp_dict.get(key, {})
            if isinstance(val, dict):
                return val.get(lang, val.get('tr', fallback))
            return fallback

        # Matrix Row Definitions
        # Top Header Rows
        header_rows = [
            ("ÇAP (inch)", lambda p: str(p['input_summary']['diameter_inch']), get_exp('diameter')),
            ("ÇAP (mm)", lambda p: f"{p['input_summary']['diameter_mm']:.2f}", "Boru Gerçek Dış Çapı (OD mm)" if lang == "tr" else "Actual Outside Diameter (OD mm)"),
            ("Design Basıncı", lambda p: str(p['input_summary']['design_factor_str']), get_exp('design_factor')),
            ("Et Kalınlığı (mm)", lambda p: f"{p['input_summary']['wall_thickness_mm']:.2f}", get_exp('wall_thickness')),
            ("Üretim Yöntemi", lambda p: str(p['input_summary']['manufacturing_process']), get_exp('process')),
            ("Malzeme Kalitesi", lambda p: str(p['input_summary']['material_grade']), get_exp('grade')),
            ("SMYS (Psi)", lambda p: f"{p['mechanical_properties']['smys_psi']:.2f}", get_exp('smys')),
        ]

        current_r = start_row
        row_remarks_map = {}

        for label, extractor, remark in header_rows:
            ws.merge_cells(start_row=current_r, start_column=1, end_row=current_r, end_column=4)
            cell_lbl = ws.cell(current_r, 1, label)
            cell_lbl.font = font_header_dark
            cell_lbl.fill = fill_header_main
            cell_lbl.alignment = align_center
            for c in range(1, 5):
                ws.cell(current_r, c).border = border_all
                ws.cell(current_r, c).fill = fill_header_main

            for idx, pipe in enumerate(pipes_data):
                col_idx = 5 + idx
                val = extractor(pipe)
                c_val = ws.cell(current_r, col_idx, val)
                c_val.font = font_header_bold
                c_val.alignment = align_center
                c_val.border = border_all
                c_val.fill = fill_header_sub if current_r <= start_row + 3 else PatternFill(fill_type=None)

            row_remarks_map[current_r] = remark
            current_r += 1

        # 2. Chemical Analysis Block (C, Mn, P, S, Nb, V, Ti, N, CE IIW, CE Pcm)
        def _fmt(v, dec):
            if v is None or v == "":
                return "—"
            if isinstance(v, (int, float)):
                return f"{v:.{dec}f}"
            return str(v)

        chem_start_r = current_r
        chem_items = [
            ("C", "Max %", lambda p: _fmt(p['chemical_analysis'].get('C_max'), 2)),
            ("Mn", "Max %", lambda p: _fmt(p['chemical_analysis'].get('Mn_max'), 2)),
            ("P", "Max %", lambda p: _fmt(p['chemical_analysis'].get('P_max'), 3)),
            ("S", "Max %", lambda p: _fmt(p['chemical_analysis'].get('S_max'), 3)),
            ("Nb", "Min%-Max%", lambda p: str(p['chemical_analysis'].get('Nb_min_max')) if p['chemical_analysis'].get('Nb_min_max') else "—"),
            ("V", "Max %", lambda p: _fmt(p['chemical_analysis'].get('V_max'), 2)),
            ("Ti", "Max %", lambda p: _fmt(p['chemical_analysis'].get('Ti_max'), 2)),
            ("N", "Max %", lambda p: _fmt(p['chemical_analysis'].get('N_max'), 3)),
            ("CE (IIW)", "Max", lambda p: _fmt(p['chemical_analysis'].get('CE_IIW_max'), 2)),
            ("CE (Pcm)", "Max", lambda p: _fmt(p['chemical_analysis'].get('CE_Pcm_max'), 2))
        ]

        for elem, limit_type, ext in chem_items:
            c_elem = ws.cell(current_r, 2, elem)
            c_elem.font = font_header_dark
            c_elem.alignment = align_center
            c_elem.fill = fill_section

            ws.merge_cells(start_row=current_r, start_column=3, end_row=current_r, end_column=4)
            c_type = ws.cell(current_r, 3, limit_type)
            c_type.font = font_header_dark
            c_type.alignment = align_center
            c_type.fill = fill_section

            for c in range(2, 5):
                ws.cell(current_r, c).border = border_all
                ws.cell(current_r, c).fill = fill_section

            for idx, pipe in enumerate(pipes_data):
                col_idx = 5 + idx
                val = ext(pipe)
                c_val = ws.cell(current_r, col_idx, val)
                c_val.font = font_regular
                c_val.alignment = align_center
                c_val.border = border_all

            row_remarks_map[current_r] = get_exp('chemical', 'API 5L Tablo 5 / BOTAŞ Çizelge 2 Kimyasal Sınırları')
            current_r += 1

        chem_end_r = current_r - 1
        ws.merge_cells(start_row=chem_start_r, start_column=1, end_row=chem_end_r, end_column=1)
        c_chem = ws.cell(chem_start_r, 1, "Kimyasal Analiz" if lang == "tr" else "Chemical Analysis")
        c_chem.font = font_header_dark
        c_chem.alignment = align_center
        c_chem.fill = fill_section
        for r in range(chem_start_r, chem_end_r + 1):
            ws.cell(r, 1).border = border_all
            ws.cell(r, 1).fill = fill_section

        # 3. Wall Thickness Tolerances
        thk_tol_start = current_r
        thk_tol_items = [
            ("Min. (mm)", lambda p: f"{p['wall_thickness_tolerance']['min_mm']:.2f}"),
            ("Max. (mm)", lambda p: f"{p['wall_thickness_tolerance']['max_mm']:.2f}")
        ]

        for lbl, ext in thk_tol_items:
            ws.merge_cells(start_row=current_r, start_column=2, end_row=current_r, end_column=4)
            c_lbl = ws.cell(current_r, 2, lbl)
            c_lbl.font = font_header_dark
            c_lbl.alignment = align_center
            c_lbl.fill = fill_section
            for c in range(2, 5):
                ws.cell(current_r, c).border = border_all
                ws.cell(current_r, c).fill = fill_section

            for idx, pipe in enumerate(pipes_data):
                col_idx = 5 + idx
                val = ext(pipe)
                c_val = ws.cell(current_r, col_idx, val)
                c_val.font = font_regular
                c_val.alignment = align_center
                c_val.border = border_all

            row_remarks_map[current_r] = get_exp('wall_thickness_tol', 'API 5L Tablo 11 / BOTAŞ Çizelge 5 Et Kalınlığı Toleransı')
            current_r += 1

        thk_tol_end = current_r - 1
        ws.merge_cells(start_row=thk_tol_start, start_column=1, end_row=thk_tol_end, end_column=1)
        c_thk = ws.cell(thk_tol_start, 1, "Et Kalınlığı\nToleransı" if lang == "tr" else "Wall Thickness\nTolerance")
        c_thk.font = font_header_dark
        c_thk.alignment = align_center
        c_thk.fill = fill_section
        for r in range(thk_tol_start, thk_tol_end + 1):
            ws.cell(r, 1).border = border_all
            ws.cell(r, 1).fill = fill_section

        def _fmt(v, decimals=2, suffix=""):
            if v is None:
                return ""
            if isinstance(v, (int, float)):
                return f"{v:.{decimals}f}{suffix}"
            return f"{v}{suffix}" if suffix and not str(v).endswith(suffix) else str(v)

        # 4. Remaining Inspection Rows (Full width label in Cols 1-4)
        inspection_rows = [
            ("Boru Çap Toleransı - Boru Ucu Max (mm)", lambda p: _fmt(p['dimensional_tolerances']['diameter_end_max_mm']), get_exp('diameter_tol')),
            ("Boru Çap Toleransı - Boru Ucu Min (mm)", lambda p: _fmt(p['dimensional_tolerances']['diameter_end_min_mm']), get_exp('diameter_tol')),
            ("Boru Çap Toleransı - Boru Gövdesi Max (mm)", lambda p: _fmt(p['dimensional_tolerances']['diameter_body_max_mm']), get_exp('diameter_tol')),
            ("Boru Çap Toleransı - Boru Gövdesi Min (mm)", lambda p: _fmt(p['dimensional_tolerances']['diameter_body_min_mm']), get_exp('diameter_tol')),
            ("Boru Çevre Toleransı - Boru Ucu Max (mm)", lambda p: str(p['dimensional_tolerances']['circ_end_max_mm']), get_exp('circumference_tol')),
            ("Boru Çevre Toleransı - Boru Ucu Min (mm)", lambda p: str(p['dimensional_tolerances']['circ_end_min_mm']), get_exp('circumference_tol')),
            ("Boru Çevre Toleransı - Boru Gövdesi Max (mm)", lambda p: str(p['dimensional_tolerances']['circ_body_max_mm']), get_exp('circumference_tol')),
            ("Boru Çevre Toleransı - Boru Gövdesi Min (mm)", lambda p: str(p['dimensional_tolerances']['circ_body_min_mm']), get_exp('circumference_tol')),
            ("Ovalite - Boru Ucu (mm)", lambda p: str(p['dimensional_tolerances']['ovality_end_mm']), get_exp('ovality')),
            ("Ovalite - Boru Gövdesi (mm)", lambda p: str(p['dimensional_tolerances']['ovality_body_mm']), get_exp('ovality')),
            ("Yield Min. (Psi-Mpa)", lambda p: f"{_fmt(p['mechanical_properties']['yield_min_psi'])} / {_fmt(p['mechanical_properties']['yield_min_mpa'])}", get_exp('yield_tensile')),
            ("Yield Max. (Psi-Mpa)", lambda p: f"{_fmt(p['mechanical_properties']['yield_max_psi'])} / {_fmt(p['mechanical_properties']['yield_max_mpa'])}", get_exp('yield_tensile')),
            ("Tensile Min (Psi-Mpa)", lambda p: f"{_fmt(p['mechanical_properties']['tensile_min_psi'])} / {_fmt(p['mechanical_properties']['tensile_min_mpa'])}", get_exp('yield_tensile')),
            ("Tensile Max (Psi-Mpa)", lambda p: f"{_fmt(p['mechanical_properties']['tensile_max_psi'])} / {_fmt(p['mechanical_properties']['tensile_max_mpa'])}", get_exp('yield_tensile')),
            ("Akma / Çekme Oranı Max. (Y/T)", lambda p: _fmt(p['mechanical_properties']['yield_to_tensile_ratio_max']), get_exp('yt_ratio')),
            ("Hydro Test Basıncı Max. (Bar)", lambda p: _fmt(p['hydrostatic_test']['hydro_test_max_bar']), get_exp('hydro_test')),
            ("Hydro Test Basıncı Min. (Bar)", lambda p: _fmt(p['hydrostatic_test']['hydro_test_min_bar']), "P_max - 2.0 Bar (Fabrika Test Alt Sınırı)"),
            ("API 5L Standart Test Pressure (Bar)", lambda p: _fmt(p['hydrostatic_test']['api_5l_std_test_bar']), get_exp('api_std_test')),
            ("Minimum Uzama (% e) - Malzeme", lambda p: _fmt(p['toughness_and_tests']['elongation_mat_min_percent'], suffix="%"), get_exp('elongation')),
            ("Minimum Uzama (% e) - Kaynak", lambda p: _fmt(p['toughness_and_tests']['elongation_weld_min_percent'], suffix="%"), "Kaynak Dikişi Min. %10 Uzama"),
            ("Çentik Darbe (J) - Malzeme", lambda p: _fmt(p['toughness_and_tests']['notch_impact_mat_j'], suffix=" J"), get_exp('cvn')),
            ("Çentik Darbe (J) - Kaynak", lambda p: _fmt(p['toughness_and_tests']['notch_impact_weld_j'], suffix=" J"), get_exp('cvn')),
            ("Radial Offset Max. (mm)", lambda p: _fmt(p['weld_and_geometry']['radial_offset_max_mm']), get_exp('radial_offset')),
            ("Kaynak Yüksekliği - İç / Dış (mm)", lambda p: f"{_fmt(p['weld_and_geometry']['weld_height_inside_mm'])} / {_fmt(p['weld_and_geometry']['weld_height_outside_mm'])}", get_exp('weld_height')),
            ("Misalignment (mm)", lambda p: _fmt(p['weld_and_geometry']['misalignment_max_mm']), get_exp('misalignment')),
            ("Artık Gerilme Testi Max (mm)", lambda p: _fmt(p['toughness_and_tests']['residual_stress_max_mm']), get_exp('residual_stress')),
            ("Yırtılma Testi (DWTT)", lambda p: str(p['toughness_and_tests']['dwtt_test']), get_exp('dwtt')),
            ("Sertlik TESTİ", lambda p: str(p['toughness_and_tests']['hardness_test_max']), get_exp('hardness')),
            ("Mandrel Çapı / Çene Açıklığı (mm)", lambda p: f"{_fmt(p['toughness_and_tests']['mandrel_dia_max_mm'])} / {_fmt(p['toughness_and_tests']['jaw_opening_max_mm'])}", get_exp('mandrel_jaw')),
            ("FLATTENING - Kaynak / Çatlak", lambda p: f"{_fmt(p['flattening']['weld_opening_height_mm'])} / {_fmt(p['flattening']['material_crack_height_mm'])}", get_exp('flattening')),
            ("Ağırlık Nominal (Kg/m)", lambda p: f"{_fmt(p['weights_and_safety']['weight_nominal_kg_m'])} ({_fmt(p['weights_and_safety']['weight_min_kg_m'])} - {_fmt(p['weights_and_safety']['weight_max_kg_m'])})", get_exp('weight')),
            ("Operating pressure / SMYS", lambda p: str(p['weights_and_safety']['operating_press_over_smys_percent']), "İşletme Gerilmesi / SMYS Oranı"),
            ("841.1.2 Fracture Control and Arrest", lambda p: str(p['weights_and_safety']['fracture_control_asme_841_1_2']), get_exp('fracture_control')),
            ("D/t & Alternatif Basınç Dizayn", lambda p: f"D/t={_fmt(p['weights_and_safety']['d_over_t'])} ({_fmt(p['weights_and_safety']['design_formula_asme_841_1_1'])})", get_exp('thick_wall_alt')),
        ]

        for lbl, ext, remark in inspection_rows:
            ws.merge_cells(start_row=current_r, start_column=1, end_row=current_r, end_column=4)
            c_lbl = ws.cell(current_r, 1, lbl)
            c_lbl.font = font_header_dark
            c_lbl.alignment = align_left
            for c in range(1, 5):
                ws.cell(current_r, c).border = border_all
                if current_r % 2 == 0:
                    ws.cell(current_r, c).fill = fill_zebra

            for idx, pipe in enumerate(pipes_data):
                col_idx = 5 + idx
                val = ext(pipe)
                c_val = ws.cell(current_r, col_idx, val)
                c_val.font = font_regular
                c_val.alignment = align_center
                c_val.border = border_all
                if current_r % 2 == 0:
                    c_val.fill = fill_zebra

            row_remarks_map[current_r] = remark
            current_r += 1

        # Set column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 16

        for idx in range(num_pipes):
            col_letter = get_column_letter(5 + idx)
            ws.column_dimensions[col_letter].width = 18

        # Set Remarks column width & fill remarks for ALL 40+ rows
        remarks_col_idx = 5 + num_pipes
        remarks_col_letter = get_column_letter(remarks_col_idx)
        ws.column_dimensions[remarks_col_letter].width = 40

        cell_rem_head = ws.cell(start_row - 1, remarks_col_idx, "Standart Referansı & Açıklama" if lang == "tr" else "Standard Reference & Remarks")
        cell_rem_head.font = font_header_dark
        cell_rem_head.fill = fill_header_main
        cell_rem_head.alignment = align_center
        cell_rem_head.border = border_all

        for r_idx, rem_text in row_remarks_map.items():
            c_exp = ws.cell(r_idx, remarks_col_idx, rem_text)
            c_exp.font = font_sub
            c_exp.fill = fill_zebra
            c_exp.alignment = align_left
            c_exp.border = border_all

        # Disclaimer & Copyright Footer in Excel
        disclaimer_row = current_r + 2
        total_cols = remarks_col_idx

        # Copyright Note
        ws.merge_cells(start_row=disclaimer_row, start_column=1, end_row=disclaimer_row, end_column=total_cols)
        c_copy = ws.cell(disclaimer_row, 1, "© 2026 API 5L PSL2 & BOTAŞ Pipe QA/QC & Design Suite. Tüm Hakları Saklıdır.")
        c_copy.font = font_header_bold
        c_copy.alignment = align_left

        # Engineering Disclaimer Note
        ws.merge_cells(start_row=disclaimer_row + 1, start_column=1, end_row=disclaimer_row + 2, end_column=total_cols)
        disclaimer_text = (
            "Yasal Uyarı / Disclaimer: Bu hesaplama ve kalite güvence tablosu mühendislik ön tasarımı ve fabrika kabul denetimleri amacıyla üretilmiştir. "
            "Projelerdeki nihai imalat ve satın alma kararlarında yürürlükteki resmi standartlar ve lisanslı baş mühendisin onayı esastır."
            if lang == "tr" else
            "Disclaimer: This calculation and QA/QC matrix is generated for engineering pre-design and factory acceptance inspection purposes. "
            "Official standards and licensed chief engineer approvals govern final manufacturing and procurement decisions."
        )
        c_disc = ws.cell(disclaimer_row + 1, 1, disclaimer_text)
        c_disc.font = font_sub
        c_disc.fill = fill_disclaimer
        c_disc.alignment = align_left

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_itp_audit_report(
        audit_data: Dict[str, Any],
        lang: str = "tr"
    ) -> io.BytesIO:
        """
        Exports the ITP comparison and audit findings to a styled Excel report.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ITP Denetim ve Sapma Raporu"
        ws.views.sheetView[0].showGridLines = True

        font_title = Font(name="Arial", size=13, bold=True, color="1E3A8A")
        font_header = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        font_bold = Font(name="Arial", size=9, bold=True)
        font_regular = Font(name="Arial", size=9)
        font_small = Font(name="Arial", size=8, italic=True, color="475569")

        fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_pass = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        fill_warn = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        fill_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        border_thin = Side(border_style="thin", color="CBD5E1")
        border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        pipe = audit_data.get("pipe_summary", {})
        kpi = audit_data.get("kpi", {})

        std_ed_title = pipe.get("standard_edition", "API Spec 5L 47. Baskı")
        # Header Title
        ws.merge_cells("A1:J1")
        ws.cell(1, 1, f"{std_ed_title} - ITP AKILLI DENETİM VE SAPMA RAPORU").font = font_title
        ws.cell(1, 1).alignment = align_left

        # Pipe Metadata Box
        ws.merge_cells("A2:J2")
        d_inch_str = pipe.get("diameter_inch", "48in")
        d_mm_str = pipe.get("diameter_mm", "1219")
        wt_mm_str = pipe.get("wall_thickness_mm", "14.3")
        grade_str = pipe.get("material_grade", "X65")
        psl_str = pipe.get("psl_level", "PSL2")
        process_str = pipe.get("manufacturing_process", "SAWH")
        score_val = kpi.get("compliance_score_percent", 0.0)
        bare_s = kpi.get("bare_pipe_score_percent")
        coat_s = kpi.get("coating_score_percent")
        verdict_str = kpi.get("overall_verdict", "N/A")

        scores_summary = f"Genel Uyum: %{score_val}"
        if bare_s is not None:
            scores_summary += f" | Çıplak Boru: %{bare_s}"
        if coat_s is not None:
            scores_summary += f" | 3LPE Kaplama: %{coat_s}"

        meta_str = (
            f"Boru Özellikleri: {d_inch_str} ({d_mm_str} mm) x {wt_mm_str} mm | "
            f"Kalite: {grade_str} {psl_str} | Üretim: {process_str} | "
            f"{scores_summary} ({verdict_str})"
        )
        ws.cell(2, 1, meta_str).font = font_small
        ws.cell(2, 1).alignment = align_left

        # Table Column Headers
        headers = [
            "Muayene / Test Adı",
            "Kategori / Disiplin",
            "Boru Sütunu Hesaplanan Hedef Değer",
            "NDT Metot Standardı",
            "İmalatçı ITP Frekansı",
            "API 5L / BOTAŞ Şartname Frekansı",
            "İmalatçı ITP Kabul Kriteri",
            "Standart Kabul Kriteri (Limit Değer)",
            "Şahitlik Noktaları (C/W/H)",
            "Denetim Durumu & Bulgular",
        ]
        
        ws.row_dimensions[4].height = 28
        for col_idx, text in enumerate(headers, 1):
            c = ws.cell(4, col_idx, text)
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_center
            c.border = border_all

        # Populate Rows
        current_r = 5
        for row in audit_data.get("audit_rows", []):
            ws.row_dimensions[current_r].height = 36
            status = row.get("status", "COMPLIANT")
            ip = row.get("inspection_points", {})
            ip_str = f"Üretici: {ip.get('mfg', 'C')} | TPI: {ip.get('tpi', 'W')} | İdare: {ip.get('client', 'W')}" if isinstance(ip, dict) and ip else "C / W / H"
            
            c_name = ws.cell(current_r, 1, row.get("test_name", "—"))
            c_cat = ws.cell(current_r, 2, row.get("category", "—"))
            c_target = ws.cell(current_r, 3, row.get("calculated_target", "—"))
            c_ndt = ws.cell(current_r, 4, row.get("ndt_method_standard", "—"))
            c_up_f = ws.cell(current_r, 5, row.get("uploaded_frequency", "—"))
            c_st_f = ws.cell(current_r, 6, row.get("standard_frequency", "—"))
            c_up_c = ws.cell(current_r, 7, row.get("uploaded_criteria", "—"))
            c_st_c = ws.cell(current_r, 8, row.get("standard_criteria", "—"))
            c_ip = ws.cell(current_r, 9, ip_str)
            c_rem = ws.cell(current_r, 10, row.get("audit_remarks", "—"))

            status_fill = fill_pass if status == "COMPLIANT" else (fill_warn if status == "MORE_STRINGENT" else fill_fail)

            for cell in (c_name, c_cat, c_target, c_ndt, c_up_f, c_st_f, c_up_c, c_st_c, c_ip, c_rem):
                cell.font = font_regular
                cell.border = border_all
                cell.alignment = align_left

            c_target.font = font_bold
            c_ip.alignment = align_center
            c_rem.fill = status_fill
            c_rem.font = font_bold if status != "COMPLIANT" else font_regular
            c_name.fill = fill_zebra if current_r % 2 == 0 else PatternFill(fill_type=None)

            current_r += 1

        # Column Widths & Freeze Panes & Auto Filter
        widths = [26, 16, 28, 22, 22, 26, 26, 30, 24, 45]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        ws.freeze_panes = "A5"
        if current_r > 5:
            ws.auto_filter.ref = f"A4:J{current_r - 1}"

        # Sheet 2: Findings & Non-Compliances
        findings = audit_data.get("findings", [])
        if findings:
            ws_f = wb.create_sheet(title="Bulgular ve Sapmalar")
            ws_f.views.sheetView[0].showGridLines = True
            ws_f.merge_cells("A1:E1")
            ws_f.cell(1, 1, "ITP DENETİMİ SAPMA VE UYARILAR LİSTESİ").font = font_title
            ws_f.cell(1, 1).alignment = align_left

            f_headers = ["Önem Seviyesi", "Madde / Test Adı", "Kusur Türü", "Şartname / Madde Ref.", "Denetçi Açıklaması"]
            ws_f.row_dimensions[3].height = 26
            for col_idx, text in enumerate(f_headers, 1):
                c = ws_f.cell(3, col_idx, text)
                c.font = font_header
                c.fill = fill_header
                c.alignment = align_center
                c.border = border_all

            f_row = 4
            for find in findings:
                ws_f.row_dimensions[f_row].height = 30
                sev = find.get("severity", "INFO")
                sev_fill = fill_fail if sev == "CRITICAL" else fill_warn

                c_sev = ws_f.cell(f_row, 1, sev)
                c_tname = ws_f.cell(f_row, 2, find.get("test_name", "—"))
                c_type = ws_f.cell(f_row, 3, find.get("issue_type", "—"))
                c_clause = ws_f.cell(f_row, 4, find.get("clause_ref", "—"))
                c_msg = ws_f.cell(f_row, 5, find.get("message", "—"))

                for cell in (c_sev, c_tname, c_type, c_clause, c_msg):
                    cell.font = font_regular
                    cell.border = border_all
                    cell.alignment = align_left

                c_sev.fill = sev_fill
                c_sev.font = font_bold
                c_sev.alignment = align_center
                f_row += 1

            f_widths = [16, 26, 22, 24, 60]
            for idx, w in enumerate(f_widths, 1):
                ws_f.column_dimensions[get_column_letter(idx)].width = w

            ws_f.freeze_panes = "A4"
            ws_f.auto_filter.ref = f"A3:E{f_row - 1}"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
